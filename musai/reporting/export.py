"""Styled XLSX evidence export — one clean, colorful workbook per group.

A professor's local audit/evidence sheet: per student, each partial's
General% / Special% / Exam% / final 0–10, plus the course total (institutional
30/30/40 weighting). Grade cells are colored on the same red→yellow→green scale
as the cockpit. The 'final' columns already include any curve + extra credit.
"""

from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from musai.db import engine
from musai.models import Course, Semester, Student, Enrollment, Partial, PartialGrade
from musai.web.format import grade_hex

# Institutional weighting of the three partials into the final course grade.
COURSE_WEIGHTS = [0.30, 0.30, 0.40]  # Parcial 1 · Parcial 2 · Examen Final Ordinario

CLAY = "B07D4B"; CLAY_DK = "855A31"; PAPER = "FAF6EF"; CREAM = "FBF8F2"
LINE = "EBE3D5"; INK = "3A362F"; MUTED = "8C8474"; WHITE = "FFFFFF"

_thin = Side(style="thin", color=LINE)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _argb(hex6: str) -> str:
    return "FF" + hex6


def _gather(course_id: int) -> dict:
    """Read everything for one group into plain dicts (no detached ORM objects)."""
    with Session(engine) as s:
        course = s.get(Course, course_id)
        if course is None:
            raise ValueError(f"No course id {course_id}")
        semester = s.get(Semester, course.semester_id)
        partials = s.exec(
            select(Partial).where(Partial.course_id == course_id).order_by(Partial.id)
        ).all()
        enrolls = s.exec(select(Enrollment).where(Enrollment.course_id == course_id)).all()
        students = [st for st in (s.get(Student, e.student_id) for e in enrolls) if st]
        students.sort(key=lambda x: x.full_name)

        pg_by = {}
        for p in partials:
            for pg in s.exec(select(PartialGrade).where(PartialGrade.partial_id == p.id)).all():
                pg_by[(pg.student_id, p.id)] = pg

        rows = []
        for st in students:
            cells, finals = {}, []
            for i, p in enumerate(partials):
                pg = pg_by.get((st.id, p.id))
                if pg:
                    comp = json.loads(pg.components_json or "{}")
                    cells[p.id] = {"gen": comp.get("general_avg"), "spec": comp.get("special"),
                                   "exam": comp.get("exam"), "final": pg.sega_value}
                    finals.append((i, pg.sega_value))
                else:
                    cells[p.id] = {"gen": None, "spec": None, "exam": None, "final": None}
            if len(finals) == len(partials) and partials:
                if len(partials) == 3:
                    total = sum(COURSE_WEIGHTS[i] * v for i, v in finals)
                else:
                    total = sum(v for _, v in finals) / len(finals)
                course_total = round(total, 1)
            else:
                course_total = None
            rows.append({"mat": st.matricula, "name": st.full_name,
                         "cells": cells, "course_total": course_total})

        return {
            "group": course.group_code, "subject": course.subject,
            "semester": semester.name if semester else "",
            "partials": [{"id": p.id, "name": p.name} for p in partials],
            "rows": rows,
        }


def build_group_workbook(course_id: int) -> Workbook:
    d = _gather(course_id)
    partials = d["partials"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    base = 3                       # #, Matrícula, Alumno
    per = 4                        # Gen, Esp, Exam, /10
    total_col = base + per * len(partials) + 1
    last_letter = get_column_letter(total_col)

    # ── title block ───────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_letter}1")
    t = ws["A1"]; t.value = "MUSAI · Resumen de calificaciones"
    t.font = Font(name="Calibri", size=16, bold=True, color=_argb(CLAY_DK))
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_letter}2")
    sub = ws["A2"]
    sub.value = (f"{d['group']} · {d['subject']} · {d['semester']}      "
                 f"Generado {datetime.now():%Y-%m-%d %H:%M}")
    sub.font = Font(size=10, color=_argb(MUTED))

    # ── header rows (4 = partial groups, 5 = sub-columns) ─────────────────────
    hdr_fill = PatternFill("solid", fgColor=_argb(CLAY))
    sub_fill = PatternFill("solid", fgColor=_argb(CREAM))
    white_bold = Font(bold=True, color=_argb(WHITE), size=10)
    muted_bold = Font(bold=True, color=_argb(CLAY_DK), size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    GH, SH = 4, 5  # group-header row, sub-header row
    # left identity headers span both header rows
    for col, label in ((1, "#"), (2, "Matrícula"), (3, "Alumno")):
        ws.merge_cells(start_row=GH, start_column=col, end_row=SH, end_column=col)
        c = ws.cell(GH, col, label); c.fill = hdr_fill; c.font = white_bold
        c.alignment = center; c.border = BORDER
        ws.cell(SH, col).border = BORDER

    for i, p in enumerate(partials):
        c0 = base + i * per + 1
        ws.merge_cells(start_row=GH, start_column=c0, end_row=GH, end_column=c0 + per - 1)
        gc = ws.cell(GH, c0, p["name"]); gc.fill = hdr_fill; gc.font = white_bold
        gc.alignment = center; gc.border = BORDER
        for j, lab in enumerate(("Gen %", "Esp %", "Exam %", "/10")):
            sc = ws.cell(SH, c0 + j, lab); sc.fill = sub_fill; sc.font = muted_bold
            sc.alignment = center; sc.border = BORDER

    # course total header (spans both rows)
    ws.merge_cells(start_row=GH, start_column=total_col, end_row=SH, end_column=total_col)
    tc = ws.cell(GH, total_col, "Curso /10"); tc.fill = hdr_fill; tc.font = white_bold
    tc.alignment = center; tc.border = BORDER
    ws.cell(SH, total_col).border = BORDER

    # ── data rows ─────────────────────────────────────────────────────────────
    data_start = SH + 1
    for ri, row in enumerate(d["rows"]):
        r = data_start + ri
        band = PatternFill("solid", fgColor=_argb(CREAM)) if ri % 2 else None
        idx = ws.cell(r, 1, ri + 1); idx.font = Font(size=9, color=_argb(MUTED))
        mat = ws.cell(r, 2, row["mat"]); mat.font = Font(size=10, color=_argb(MUTED))
        nm = ws.cell(r, 3, row["name"]); nm.font = Font(size=10, color=_argb(INK))
        for cc in (idx, mat, nm):
            cc.border = BORDER
            if band:
                cc.fill = band
            cc.alignment = Alignment(horizontal="left" if cc is nm else "center", vertical="center")

        for i, p in enumerate(partials):
            c0 = base + i * per + 1
            cell = row["cells"][p["id"]]
            for j, key in enumerate(("gen", "spec", "exam")):
                v = cell[key]
                cx = ws.cell(r, c0 + j, round(v) if v is not None else "—")
                cx.font = Font(size=10, color=_argb(MUTED))
                cx.alignment = Alignment(horizontal="center", vertical="center")
                cx.border = BORDER
                if band:
                    cx.fill = band
            _grade_cell(ws, r, c0 + 3, cell["final"])
        _grade_cell(ws, r, total_col, row["course_total"], bold=True)

    # ── sizing, freeze, footer ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 4.5
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 30
    for col in range(base + 1, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8.5
    ws.freeze_panes = ws.cell(data_start, 4)

    foot = data_start + len(d["rows"]) + 1
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=total_col)
    fc = ws.cell(foot, 1,
                 "Final /10 = nota exacta + curva + crédito extra (transparente y auditado). "
                 "Curso = P1·30% + P2·30% + Ordinario·40%. Aprobatoria 7.0.")
    fc.font = Font(size=8, italic=True, color=_argb(MUTED))
    return wb


def _grade_cell(ws, r, c, value, bold: bool = False):
    cell = ws.cell(r, c, value if value is not None else "—")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    if value is None:
        cell.font = Font(size=10, color=_argb(MUTED))
        return
    hx = grade_hex(value)
    cell.fill = PatternFill("solid", fgColor=_argb(hx["fill"]))
    cell.font = Font(size=10, bold=True, color=_argb(hx["font"]))
    cell.number_format = "0.0"


def workbook_bytes(course_id: int) -> bytes:
    buf = io.BytesIO()
    build_group_workbook(course_id).save(buf)
    return buf.getvalue()


def save_group_workbook(course_id: int, path: str | Path) -> Path:
    p = Path(path)
    build_group_workbook(course_id).save(p)
    return p
